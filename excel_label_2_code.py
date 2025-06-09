from sentence_transformers import SentenceTransformer, util
import torch
import csv
import re
from unidecode import unidecode
import os
import openpyxl
import json

#setup parameters
MODEL_NAME_REMOTE = 'paraphrase-multilingual-MiniLM-L12-v2' #this is the multilingual model (example)
LOCAL_MODEL_DIRECTORY = "./model_local" #thi is the local path for model downloading (for offline use)
model = None
dataset_labels_original = []
dataset_embeddings = None
dataset_map_original_to_code = {}

#function to load, from a csv file, the reference dataset containing two columns: label - code
def load_ref_dataset(file_path):
    dataset = {}
    try:
        with open(file_path, mode='r', encoding='utf-8', newline='') as file_csv:
            csv_reader = csv.reader(file_csv)
            try:
                header = next(csv_reader)
            except StopIteration:
                pass
            for i, row in enumerate(csv_reader):
                if len(row) == 2:
                    label = row[0].strip()
                    code = row[1].strip()
                    if label and code:
                        dataset[label] = code
    except FileNotFoundError:
        print(f"Error: reference CSV file not found at '{file_path}'")
    except Exception as e:
        print(f"Error while reading reference CSV file '{file_path}': {e}")
    return dataset

#cleaning function
def preprocess_text_for_embedding(text):
    if not text: return ""
    return unidecode(str(text)).strip()

#function for source dataset embedding creations
def initialize_model_and_dataset_embeddings(path_csv_file_reference):
    global model, dataset_labels_original, dataset_embeddings, dataset_map_original_to_code
    model_loaded_locally = False
    if os.path.exists(LOCAL_MODEL_DIRECTORY):
        try:
            model = SentenceTransformer(LOCAL_MODEL_DIRECTORY)
            model_loaded_locally = True
            print("Model successfully loaded from local directory.")
        except Exception as e:
            print(f"Error loading model from local: {e}. Proceeding with download.")
    if not model_loaded_locally:
        try:
            print(f"Attempt to download/upload model from Hugging Face: {MODEL_NAME_REMOTE}")
            model = SentenceTransformer(MODEL_NAME_REMOTE)
            print("Model successfully downloaded/uploaded from Hugging Face.")
            if not os.path.exists(LOCAL_MODEL_DIRECTORY): os.makedirs(LOCAL_MODEL_DIRECTORY)
            model.save(LOCAL_MODEL_DIRECTORY)
            print(f"Model saved locally in '{LOCAL_MODEL_DIRECTORY}'.")
        except Exception as e:
            print(f"Critical error in downloading/uploading SentenceTransformer model: {e}")
            return False

    dataset_map_original_to_code = load_ref_dataset(path_csv_file_reference)
    if not dataset_map_original_to_code:
        print("Reference dataset not loaded or empty.")
        return False

    dataset_labels_original = list(dataset_map_original_to_code.keys())
    processed_labels = [preprocess_text_for_embedding(label) for label in dataset_labels_original]

    if not processed_labels:
        print("No valid label in the reference dataset to calculate embeddings.")
        dataset_embeddings = torch.empty(0)
        return True

    try:
        print("Calculating embeddings for the reference dataset...")
        dataset_embeddings = model.encode(processed_labels, convert_to_tensor=True, show_progress_bar=False)
        print("Calculated Embeddings.")
        return True
    except Exception as e:
        print(f"Error during encoding of reference dataset: {e}")
        return False

#function to return the most similar match
def trova_codice_con_embeddings(label_utente_originale, threshold_similarity=0.7):
    global model, dataset_labels_original, dataset_embeddings, dataset_map_original_to_code
    if not model or dataset_embeddings is None or dataset_embeddings.nelement() == 0:
        return None, None, 0
    if not label_utente_originale: return None, None, 0

    processed_label = preprocess_text_for_embedding(label_utente_originale)
    if not processed_label: return None, None, 0

    try:
        query_embedding = model.encode(processed_label, convert_to_tensor=True)
    except Exception as e:
        print(f"Errore encoding query '{processed_label}' (orig: '{label_utente_originale}'): {e}")
        return None, None, 0

    hits = util.semantic_search(query_embedding, dataset_embeddings, top_k=1)
    if not hits or not hits[0]: return None, None, 0

    best_hit = hits[0][0]
    score, idx = best_hit['score'], best_hit['corpus_id']

    if idx >= len(dataset_labels_original):
        print(f"Error: index {idx} out of bounds for dataset_labels_original (len: {len(dataset_labels_original)})")
        return None, None, 0

    matched_label_in_dataset = dataset_labels_original[idx]

    if score >= threshold_similarity:
        return dataset_map_original_to_code[matched_label_in_dataset], matched_label_in_dataset, score
    return None, matched_label_in_dataset, score

#support function to find if a string is a number or not
def is_number(s):
    if s is None: return False
    try:
        clean_s = str(s).strip()
        if not clean_s: return False
        if '.' in clean_s and ',' in clean_s:
            if clean_s.rfind('.') > clean_s.rfind(','):
                clean_s = clean_s.replace(',', '')
            else:
                clean_s = clean_s.replace('.', '').replace(',', '.')
        elif ',' in clean_s:
             clean_s = clean_s.replace(',', '.')
        float(clean_s)
        return True
    except ValueError: return False

#function for excel nomalization: every merged cell must be replicated (exploded)
def normalize_excel_to_list_of_lists(excel_filepath, sheet_name_or_index=0):
    try:
        workbook = openpyxl.load_workbook(excel_filepath, data_only=True)
        if isinstance(sheet_name_or_index, int):
            sheet = workbook.worksheets[sheet_name_or_index]
        else:
            sheet = workbook[sheet_name_or_index]
    except FileNotFoundError:
        print(f"Error: Excel file not found at '{excel_filepath}'")
        return []
    except Exception as e: # Cattura più generica per altri errori di openpyxl
        print(f"Error while opening or reading Excel file '{excel_filepath}': {e}")
        return []

    if not hasattr(sheet, 'max_row') or not hasattr(sheet, 'max_column') or \
       sheet.max_row == 0 or sheet.max_column == 0:
        print(f"Warning: the sheet '{sheet.title if hasattr(sheet, 'title') else 'N/A'}' appears to be empty or invalid.")
        return []

    data = [[cell.value for cell in row] for row in sheet.iter_rows(max_row=sheet.max_row, max_col=sheet.max_column)]

    for merged_cell_range in sheet.merged_cells.ranges:
        r_min, c_min, r_max, c_max = merged_cell_range.min_row, merged_cell_range.min_col, merged_cell_range.max_row, merged_cell_range.max_col
        top_left_cell_value = sheet.cell(row=r_min, column=c_min).value
        for r_idx in range(r_min - 1, r_max):
            for c_idx in range(c_min - 1, c_max):
                if r_idx < len(data) and c_idx < len(data[r_idx]):
                    data[r_idx][c_idx] = top_left_cell_value

    string_data = []
    for row_idx, row in enumerate(data):
        current_row_len = len(row)
        if current_row_len < sheet.max_column:
            row.extend([""] * (sheet.max_column - current_row_len))
        string_data.append([str(cell_value).strip() if cell_value is not None else "" for cell_value in row])

    print(f"Excel file '{excel_filepath}' normalized successfully.")
    return string_data

#function to process the source normalized excel and find semantical similarity
def process_data_grid(input_data_grid, threshold_similarity=0.7):
    global dataset_map_original_to_code

    if not input_data_grid:
        print("No data to process.")
        return

    num_rows = len(input_data_grid)
    if num_rows == 0:
        print("The input grid is empty.")
        return
    num_cols = 0
    for row in input_data_grid:
        if row is not None:
            num_cols = max(num_cols, len(row))

    if num_cols == 0:
        print("The input grid has no columns.")
        return

    output_data = []
    for r_idx in range(num_rows):
        row_content = input_data_grid[r_idx] if input_data_grid[r_idx] is not None else []
        output_data.append(list(row_content) + [''] * (num_cols - len(row_content)))


    for r in range(num_rows):
        for c in range(num_cols):
            if r >= len(input_data_grid) or input_data_grid[r] is None or c >= len(input_data_grid[r]):
                output_data[r][c] = ""
                continue

            original_cell_value = input_data_grid[r][c].strip()
            codice_trovato = None

            if not original_cell_value:
                output_data[r][c] = ""
                continue

            #if the cell contain a number, do not try to translate label -> code
            if is_number(original_cell_value):
                output_data[r][c] = original_cell_value
                continue

            codice_emb, _, score = trova_codice_con_embeddings(
                original_cell_value,
                threshold_similarity=threshold_similarity
            )

            if codice_emb:
                codice_trovato = codice_emb

            if codice_trovato:
                output_data[r][c] = codice_trovato
            else:
                output_data[r][c] = original_cell_value

            print(f"...for {original_cell_value} finded code {output_data[r][c]} whit score {score}")

    return output_data

#use example
if __name__ == "__main__":
    file_name_dataset_reference = "dataset.csv"
    input_excel_file_name = "test.xlsx"

    if initialize_model_and_dataset_embeddings(file_name_dataset_reference): 
        normalized_data_from_excel = normalize_excel_to_list_of_lists(input_excel_file_name)

        if normalized_data_from_excel:
            converted_output = process_data_grid(normalized_data_from_excel, threshold_similarity=0.65)
            print("Translated:")
            print(converted_output)
        else:
            print("Error: unable to translate label to code.")
    else:
        print("Execution aborted due to initialization errors.")